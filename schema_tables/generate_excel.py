import csv
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────────────────────
TABLE_COLOURS = {
    "Employee":         "1F3864",   # deep navy
    "Customer":         "1B5E20",   # forest green
    "Ticket":           "4A1580",   # deep purple
    "TicketFollowUp":   "B45309",   # amber-brown
    "TicketAssignment": "7B1FA2",   # violet
    "TicketHistory":    "C62828",   # deep red
}
HEADER_BG  = "2C3E50"
HEADER_FG  = "FFFFFF"
ROW_ALT    = "F4F6F9"
ROW_NORMAL = "FFFFFF"
BORDER_COL = "D0D5DD"

thin = Side(border_style="thin", color=BORDER_COL)
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_border(cell):
    cell.border = cell_border

def style_header(cell, bg=HEADER_BG, fg=HEADER_FG):
    cell.fill   = PatternFill("solid", fgColor=bg)
    cell.font   = Font(bold=True, color=fg, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    apply_border(cell)

def style_data(cell, alt=False, bold=False):
    cell.fill  = PatternFill("solid", fgColor=ROW_ALT if alt else ROW_NORMAL)
    cell.font  = Font(bold=bold, size=9)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    apply_border(cell)


# ── Load CSV ─────────────────────────────────────────────────────────────────
rows_by_table: dict[str, list] = {}
with open("db_schema.csv", newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        t = row["Table Name"]
        rows_by_table.setdefault(t, []).append(row)

COLUMNS = [
    "Column Name", "Data Type",
    "Constraints / Attributes", "Description", "Example Data"
]
COL_WIDTHS = [22, 14, 38, 52, 38]

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default sheet

# ── Overview sheet ────────────────────────────────────────────────────────────
ov = wb.create_sheet("Overview")
ov.sheet_view.showGridLines = False
ov.column_dimensions["A"].width = 26
ov.column_dimensions["B"].width = 70

ov.row_dimensions[1].height = 32
c = ov.cell(1, 1, "Safeway EMS — Database Schema")
c.font  = Font(bold=True, size=14, color="1F3864")
c.alignment = Alignment(vertical="center")
ov.merge_cells("A1:B1")

ov.cell(2, 1, "Generated from: prisma/schema.prisma").font = Font(italic=True, size=9, color="6B7280")
ov.cell(3, 1, "").value = None

headers = ["Table / Model", "Purpose"]
for ci, h in enumerate(headers, 1):
    style_header(ov.cell(4, ci))
    ov.cell(4, ci).value = h

PURPOSE = {
    "Employee":         "Stores Admin and Technician login accounts with role-based access.",
    "Customer":         "Stores client/company information linked to tickets.",
    "Ticket":           "Core lifecycle entity — tracks enquiries across ENQUIRY → REFILLING → SERVICES → COMPLETED stages.",
    "TicketFollowUp":   "Stores timestamped follow-up call notes linked to a ticket.",
    "TicketAssignment": "Links a technician (Employee) to a Ticket for a site visit task.",
    "TicketHistory":    "Audit trail recording every stage and status transition for a ticket.",
}

for ri, (table, purpose) in enumerate(PURPOSE.items(), 5):
    colour = TABLE_COLOURS.get(table, "374151")
    tc = ov.cell(ri, 1, table)
    tc.fill  = PatternFill("solid", fgColor=colour)
    tc.font  = Font(bold=True, color="FFFFFF", size=10)
    tc.alignment = Alignment(vertical="center")
    apply_border(tc)

    pc = ov.cell(ri, 2, purpose)
    pc.alignment = Alignment(vertical="center", wrap_text=True)
    apply_border(pc)
    ov.row_dimensions[ri].height = 22


# ── Per-table sheets ──────────────────────────────────────────────────────────
for table_name, rows in rows_by_table.items():
    ws = wb.create_sheet(table_name)
    ws.sheet_view.showGridLines = False

    colour = TABLE_COLOURS.get(table_name, "374151")

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title = ws.cell(1, 1, f"  {table_name}  —  Database Model")
    title.fill  = PatternFill("solid", fgColor=colour)
    title.font  = Font(bold=True, color="FFFFFF", size=13)
    title.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    # Column headers
    for ci, (col, width) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        c = ws.cell(2, ci, col)
        style_header(c)
    ws.row_dimensions[2].height = 20

    # Data rows
    for ri, row in enumerate(rows, 3):
        alt = (ri % 2 == 0)
        for ci, col in enumerate(COLUMNS, 1):
            val = row.get(col, "")
            c   = ws.cell(ri, ci, val)
            is_name = (ci == 1)
            style_data(c, alt=alt, bold=is_name)
        ws.row_dimensions[ri].height = 30

    # Freeze panes under header
    ws.freeze_panes = "A3"

# ── Auto-fit overview rows ────────────────────────────────────────────────────
for ri in range(5, 5 + len(PURPOSE)):
    ov.row_dimensions[ri].height = 22

wb.save("database_schema.xlsx")
print(f"[OK] database_schema.xlsx written  -- {len(rows_by_table)} table sheets + Overview")
for t, rows in rows_by_table.items():
    print(f"   {t}: {len(rows)} columns")
